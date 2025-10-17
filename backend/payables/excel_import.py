"""
Módulo para importação e exportação de contas a pagar via Excel
"""
from django.http import HttpResponse
from django.contrib import messages
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from decimal import Decimal

from .models import AccountPayable
from registrations.models import Supplier, Filial, Category, PaymentMethod


def export_template_excel(request):
    """
    Gera e exporta uma planilha modelo para importação de contas a pagar
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Contas a Pagar"

    # Estilo para cabeçalho
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Cabeçalhos
    headers = [
        "Nome da Filial",
        "CNPJ da Filial (opcional)",
        "Nome do Fornecedor",
        "CNPJ do Fornecedor (opcional)",
        "Categoria",
        "Método Pagamento",
        "Descrição",
        "Valor Original",
        "Desconto",
        "Juros",
        "Multa",
        "Valor Pago",
        "Data Emissão (DD/MM/AAAA)",
        "Data Vencimento (DD/MM/AAAA)",
        "Data Pagamento (DD/MM/AAAA)",
        "Status",
        "Nº Nota Fiscal",
        "Nº Boleto",
        "Observações",
        "É Recorrente? (SIM/NAO)",
        "Frequência Recorrência",
    ]

    # Escreve cabeçalhos
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Ajusta largura das colunas
    column_widths = [25, 20, 30, 20, 20, 20, 40, 15, 12, 12, 12, 15, 20, 20, 20, 15, 20, 20, 40, 18, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Adiciona exemplos nas primeiras 3 linhas
    exemplos = [
        [
            "Matriz",
            "12345678000190",
            "Fornecedor Exemplo LTDA",
            "98765432000100",
            "Serviços",
            "Boleto",
            "Pagamento de serviços mensais",
            "1500.00",
            "0",
            "0",
            "0",
            "0",
            "15/10/2025",
            "25/10/2025",
            "",
            "due",
            "NF-12345",
            "BOL-67890",
            "Pagamento ref. outubro/2025",
            "SIM",
            "monthly"
        ],
        [
            "Filial Centro",
            "",
            "Fornecedor Sem CNPJ",
            "",
            "Aluguel",
            "Transferência",
            "Aluguel do imóvel comercial",
            "2500.00",
            "100",
            "0",
            "0",
            "2500.00",
            "01/10/2025",
            "10/10/2025",
            "10/10/2025",
            "paid",
            "",
            "",
            "Pagamento de aluguel",
            "NAO",
            ""
        ]
    ]

    for row_num, exemplo in enumerate(exemplos, 2):
        for col_num, value in enumerate(exemplo, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Adiciona instruções em uma aba separada
    ws_instrucoes = wb.create_sheet("Instruções")
    instrucoes = [
        ["INSTRUÇÕES PARA PREENCHIMENTO DA PLANILHA"],
        [""],
        ["1. FILIAL:"],
        ["   - Nome da Filial: Nome completo da filial (OBRIGATÓRIO)"],
        ["   - CNPJ da Filial: Apenas números (14 dígitos) - OPCIONAL"],
        ["   - Se CNPJ fornecido e não existir: filial será criada"],
        ["   - Se CNPJ fornecido e já existir: usa o cadastro existente"],
        ["   - Se CNPJ vazio: busca por nome ou cria nova filial"],
        [""],
        ["2. FORNECEDOR:"],
        ["   - Nome do Fornecedor: Nome completo do fornecedor (OBRIGATÓRIO)"],
        ["   - CNPJ do Fornecedor: Apenas números (14 dígitos) - OPCIONAL"],
        ["   - Se CNPJ fornecido e não existir: fornecedor será criado"],
        ["   - Se CNPJ fornecido e já existir: usa o cadastro existente"],
        ["   - Se CNPJ vazio: busca por nome ou cria novo fornecedor"],
        [""],
        ["3. CATEGORIA e MÉTODO DE PAGAMENTO:"],
        ["   - Digite o nome exato"],
        ["   - Se não existir, será criado automaticamente"],
        [""],
        ["4. VALORES:"],
        ["   - Use ponto (.) para decimais: 1234.56"],
        ["   - Valores opcionais: Desconto, Juros, Multa, Valor Pago podem ser 0 ou vazios"],
        ["   - Valor Pago: deixe 0 ou vazio se não foi pago ainda"],
        [""],
        ["5. DATAS:"],
        ["   - Formato obrigatório: DD/MM/AAAA"],
        ["   - Exemplo: 25/10/2025"],
        ["   - Data Pagamento: deixe vazio se ainda não foi pago"],
        [""],
        ["6. STATUS:"],
        ["   - Valores aceitos (em inglês):"],
        ["     * pending - Pendente"],
        ["     * due - À Vencer (padrão se vazio)"],
        ["     * overdue - Vencida"],
        ["     * paid - Paga"],
        ["     * partially_paid - Paga Parcialmente"],
        ["     * cancelled - Cancelada"],
        [""],
        ["7. RECORRÊNCIA:"],
        ["   - É Recorrente?: Digite SIM ou NAO"],
        ["   - Frequência (se SIM): weekly, biweekly, monthly, bimonthly, quarterly, semiannual, annual"],
        [""],
        ["8. CAMPOS OBRIGATÓRIOS:"],
        ["   - Nome da Filial, Nome do Fornecedor, Categoria, Método Pagamento"],
        ["   - Descrição, Valor Original, Data Vencimento"],
        [""],
        ["9. CAMPOS OPCIONAIS:"],
        ["   - CNPJ da Filial, CNPJ do Fornecedor, Status, Valor Pago, Data Pagamento"],
        ["   - Desconto, Juros, Multa, Data Emissão, Nº NF, Nº Boleto, Observações, Recorrência"],
        [""],
        ["IMPORTANTE:"],
        ["- Não altere os cabeçalhos da primeira linha"],
        ["- Remova as linhas de exemplo antes de importar"],
        ["- Salve o arquivo no formato .xlsx"],
        ["- CNPJs são OPCIONAIS, mas se fornecidos devem ter exatamente 14 dígitos"],
        ["- Campos numéricos vazios serão tratados como 0 (zero)"],
        ["- Campos de texto vazios serão tratados como vazio (não NULL)"],
    ]

    for row_num, linha in enumerate(instrucoes, 1):
        cell = ws_instrucoes.cell(row=row_num, column=1, value=linha[0])
        if row_num == 1:
            cell.font = Font(bold=True, size=14)
        ws_instrucoes.column_dimensions['A'].width = 80

    # Prepara response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=modelo_contas_pagar_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb.save(response)
    return response


def import_excel(request, file, tenant):
    """
    Importa contas a pagar de um arquivo Excel

    Args:
        request: Request do Django
        file: Arquivo Excel enviado
        tenant: Instância do Tenant para o qual importar

    Returns:
        dict: Resultado da importação com sucesso, erros e avisos
    """
    try:
        wb = load_workbook(file)
        ws = wb.active

        resultado = {
            'sucesso': 0,
            'erros': [],
            'avisos': [],
            'criados': {
                'filiais': [],
                'fornecedores': [],
                'categorias': [],
                'metodos_pagamento': []
            }
        }

        # Pula o cabeçalho (linha 1)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Pula linhas vazias
            if not any(row):
                continue

            try:
                # Extrai dados da linha (colunas atualizadas)
                filial_nome = row[0]
                filial_cnpj = row[1] if row[1] else None  # CNPJ agora é opcional
                fornecedor_nome = row[2]
                fornecedor_cnpj = row[3] if row[3] else None  # CNPJ agora é opcional
                categoria_nome = row[4]
                metodo_pagamento_nome = row[5]
                descricao = row[6]
                valor_original = row[7]
                desconto = row[8] if row[8] not in [None, ''] else 0
                juros = row[9] if row[9] not in [None, ''] else 0
                multa = row[10] if row[10] not in [None, ''] else 0
                valor_pago = row[11] if row[11] not in [None, ''] else 0
                data_emissao = row[12] if row[12] else None
                data_vencimento = row[13]
                data_pagamento = row[14] if row[14] else None
                status = row[15] if row[15] else 'due'  # Status padrão: 'due' (À Vencer)
                nf_numero = row[16] if row[16] else ''
                boleto_numero = row[17] if row[17] else ''
                observacoes = row[18] if row[18] else ''
                is_recurring = row[19] if row[19] else None
                recurrence_frequency = row[20] if row[20] else None

                # Validações básicas - apenas campos obrigatórios
                if not all([filial_nome, fornecedor_nome, categoria_nome, metodo_pagamento_nome, descricao, valor_original, data_vencimento]):
                    resultado['erros'].append(
                        f"Linha {row_num}: Campos obrigatórios faltando. "
                        f"Verifique: Nome Filial, Nome Fornecedor, Categoria, Método Pagamento, Descrição, Valor Original e Data Vencimento"
                    )
                    continue

                # === BUSCA OU CRIA FILIAL ===
                filial = get_or_create_filial(filial_nome, filial_cnpj, tenant, resultado)

                # === BUSCA OU CRIA FORNECEDOR ===
                fornecedor = get_or_create_supplier(fornecedor_nome, fornecedor_cnpj, tenant, resultado)

                # === BUSCA OU CRIA CATEGORIA ===
                categoria = get_or_create_category(categoria_nome, tenant, resultado)

                # === BUSCA OU CRIA MÉTODO DE PAGAMENTO ===
                metodo_pagamento = get_or_create_payment_method(metodo_pagamento_nome, tenant, resultado)

                # Valida status
                status_validos = ['pending', 'due', 'overdue', 'paid', 'partially_paid', 'cancelled']
                if status.lower() not in status_validos:
                    resultado['avisos'].append(
                        f"Linha {row_num}: Status '{status}' inválido. Usando 'due' (À Vencer). "
                        f"Valores válidos: {', '.join(status_validos)}"
                    )
                    status = 'due'
                else:
                    status = status.lower()

                # Processa datas
                if isinstance(data_vencimento, str):
                    data_vencimento = datetime.strptime(data_vencimento, "%d/%m/%Y").date()
                elif hasattr(data_vencimento, 'date'):
                    # Se for datetime do Excel, converte para date
                    data_vencimento = data_vencimento.date()

                if data_emissao:
                    if isinstance(data_emissao, str):
                        data_emissao = datetime.strptime(data_emissao, "%d/%m/%Y").date()
                    elif hasattr(data_emissao, 'date'):
                        # Se for datetime do Excel, converte para date
                        data_emissao = data_emissao.date()
                else:
                    data_emissao = datetime.now().date()

                if data_pagamento:
                    if isinstance(data_pagamento, str):
                        data_pagamento = datetime.strptime(data_pagamento, "%d/%m/%Y").date()
                    elif hasattr(data_pagamento, 'date'):
                        # Se for datetime do Excel, converte para date
                        data_pagamento = data_pagamento.date()
                else:
                    data_pagamento = None

                # Processa recorrência
                is_recurring_bool = False
                if isinstance(is_recurring, str):
                    is_recurring_bool = is_recurring.upper() in ['SIM', 'YES', 'S', 'Y']

                # Cria a conta a pagar
                account = AccountPayable.objects.create(
                    tenant=tenant,
                    branch=filial,
                    supplier=fornecedor,
                    category=categoria,
                    payment_method=metodo_pagamento,
                    description=descricao,
                    original_amount=Decimal(str(valor_original)),
                    discount=Decimal(str(desconto)),
                    interest=Decimal(str(juros)),
                    fine=Decimal(str(multa)),
                    paid_amount=Decimal(str(valor_pago)),
                    issue_date=data_emissao,
                    due_date=data_vencimento,
                    payment_date=data_pagamento,
                    status=status,
                    invoice_numbers=nf_numero or '',
                    bank_slip_number=boleto_numero or '',
                    notes=observacoes or '',
                    is_recurring=is_recurring_bool,
                    recurrence_frequency=recurrence_frequency if is_recurring_bool else None,
                )

                resultado['sucesso'] += 1

            except Exception as e:
                resultado['erros'].append(f"Linha {row_num}: {str(e)}")
                continue

        return resultado

    except Exception as e:
        return {
            'sucesso': 0,
            'erros': [f"Erro ao processar arquivo: {str(e)}"],
            'avisos': [],
            'criados': {}
        }


def get_or_create_filial(nome, cnpj, tenant, resultado):
    """
    Busca ou cria uma filial.
    - Se CNPJ fornecido: busca por CNPJ (único por tenant)
    - Se CNPJ não fornecido: busca por nome (case-insensitive)
    - Se não encontrar: cria nova filial
    """
    cnpj_clean = None

    # Se CNPJ fornecido, limpa e valida
    if cnpj:
        cnpj_clean = ''.join(filter(str.isdigit, str(cnpj)))

        # Valida que o CNPJ tem 14 dígitos
        if len(cnpj_clean) != 14:
            raise ValueError(f"CNPJ inválido: {cnpj}. Deve conter exatamente 14 dígitos numéricos.")

        # Busca por CNPJ e tenant
        filial = Filial.objects.filter(tenant=tenant, cnpj=cnpj_clean).first()

        if filial:
            # Se já existe, verifica se o nome é diferente
            if filial.name.upper() != nome.upper():
                resultado['avisos'].append(
                    f"Filial com CNPJ {cnpj_clean} já existe como '{filial.name}'. "
                    f"Nome informado '{nome}' foi ignorado."
                )
            return filial
    else:
        # Se CNPJ não fornecido, busca por nome
        filial = Filial.objects.filter(tenant=tenant, name__iexact=nome).first()

        if filial:
            resultado['avisos'].append(
                f"Filial '{filial.name}' encontrada pelo nome (sem CNPJ fornecido)."
            )
            return filial

    # Cria nova filial
    # Usa um CNPJ único gerado se não for fornecido (evita UNIQUE constraint)
    if not cnpj_clean:
        import time
        # Gera um "CNPJ" único de 14 dígitos baseado em timestamp
        cnpj_temp = str(int(time.time() * 1000000))[-14:].zfill(14)
    else:
        cnpj_temp = cnpj_clean

    filial = Filial.objects.create(
        tenant=tenant,
        name=nome,
        cnpj=cnpj_temp
    )

    if cnpj_clean:
        resultado['criados']['filiais'].append(f"{nome} (CNPJ: {cnpj_clean})")
        resultado['avisos'].append(f"Filial '{nome}' (CNPJ: {cnpj_clean}) criada automaticamente")
    else:
        resultado['criados']['filiais'].append(f"{nome} (sem CNPJ)")
        resultado['avisos'].append(f"Filial '{nome}' (sem CNPJ) criada automaticamente com identificador temporário")

    return filial


def get_or_create_supplier(nome, cnpj, tenant, resultado):
    """
    Busca ou cria um fornecedor.
    - Se CNPJ fornecido: busca por CNPJ (único por tenant)
    - Se CNPJ não fornecido: busca por nome (case-insensitive)
    - Se não encontrar: cria novo fornecedor
    """
    cnpj_clean = None

    # Se CNPJ fornecido, limpa e valida
    if cnpj:
        cnpj_clean = ''.join(filter(str.isdigit, str(cnpj)))

        # Valida que o CNPJ tem 14 dígitos
        if len(cnpj_clean) != 14:
            raise ValueError(f"CNPJ do fornecedor inválido: {cnpj}. Deve conter exatamente 14 dígitos numéricos.")

        # Busca por CNPJ e tenant
        supplier = Supplier.objects.filter(tenant=tenant, cnpj=cnpj_clean).first()

        if supplier:
            # Se já existe, verifica se o nome é diferente
            if supplier.name.upper() != nome.upper():
                resultado['avisos'].append(
                    f"Fornecedor com CNPJ {cnpj_clean} já existe como '{supplier.name}'. "
                    f"Nome informado '{nome}' foi ignorado."
                )
            return supplier
    else:
        # Se CNPJ não fornecido, busca por nome
        supplier = Supplier.objects.filter(tenant=tenant, name__iexact=nome).first()

        if supplier:
            resultado['avisos'].append(
                f"Fornecedor '{supplier.name}' encontrado pelo nome (sem CNPJ fornecido)."
            )
            return supplier

    # Cria novo fornecedor
    # Usa um CNPJ único gerado se não for fornecido (evita UNIQUE constraint)
    if not cnpj_clean:
        import time
        # Gera um "CNPJ" único de 14 dígitos baseado em timestamp
        cnpj_temp = str(int(time.time() * 1000000))[-14:].zfill(14)
    else:
        cnpj_temp = cnpj_clean

    supplier = Supplier.objects.create(
        tenant=tenant,
        name=nome,
        cnpj=cnpj_temp
    )

    if cnpj_clean:
        resultado['criados']['fornecedores'].append(f"{nome} (CNPJ: {cnpj_clean})")
        resultado['avisos'].append(f"Fornecedor '{nome}' (CNPJ: {cnpj_clean}) criado automaticamente")
    else:
        resultado['criados']['fornecedores'].append(f"{nome} (sem CNPJ)")
        resultado['avisos'].append(f"Fornecedor '{nome}' (sem CNPJ) criado automaticamente com identificador temporário")

    return supplier


def get_or_create_category(nome, tenant, resultado):
    """Busca ou cria uma categoria por nome"""
    category = Category.objects.filter(tenant=tenant, name__iexact=nome).first()
    if category:
        return category

    # Cria nova categoria
    category = Category.objects.create(tenant=tenant, name=nome)
    resultado['criados']['categorias'].append(nome)
    resultado['avisos'].append(f"Categoria '{nome}' criada automaticamente")
    return category


def get_or_create_payment_method(nome, tenant, resultado):
    """Busca ou cria um método de pagamento por nome"""
    method = PaymentMethod.objects.filter(tenant=tenant, name__iexact=nome).first()
    if method:
        return method

    # Cria novo método
    method = PaymentMethod.objects.create(tenant=tenant, name=nome)
    resultado['criados']['metodos_pagamento'].append(nome)
    resultado['avisos'].append(f"Método de pagamento '{nome}' criado automaticamente")
    return method
